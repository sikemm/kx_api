#__coding__:'utf-8'
#auther:ly
from openpyxl import load_workbook

class DoExcel:
    '''该类完成从excel中读取数据，并写回测试结果'''
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name  #excel工作簿文件名或者路径
        self.sheet_name = sheet_name  #表单名

    def read_data(self):
        '''读取excel表格数据,数据格式 [{},{}]'''
        #---打开工作簿，并返回工作簿--
        wb = load_workbook(self.file_name)
        #获取表单名
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2,sheet.max_row + 1):
            row_data ={}  #存储每一行的数据
            row_data['CaseId'] = sheet.cell(i,1).value
            row_data['Module'] = sheet.cell(i,2).value
            row_data['url'] = sheet.cell(i,3).value
            row_data['Title'] = sheet.cell(i,4).value
            row_data['Method'] = sheet.cell(i, 5).value
            row_data['Params'] = sheet.cell(i, 6).value
            row_data['ExpectedResult'] = sheet.cell(i, 7).value
            row_data['Title'] = sheet.cell(i, 8).value

            test_data.append(row_data)
        wb.close()
        return test_data
    def write_data(self,row,col,value):
        '''写回测试结果到excel'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,col).value = value
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    p = DoExcel(r'F:\AutoTest\PosApi\kx_api\test_cases\api_case.xlsx','BaseInfo')
    data = p.write_data(2,8,'11111')
